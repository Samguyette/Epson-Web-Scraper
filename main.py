#!/usr/bin/env python
# File: Epson_WS_V1.py
# Name: Samuel Guyette
# Desc: Pulls data from numerous websites to compare pricing of T-series competitors.
#		Will output a .csv file to location of program.
# Other files required: ws_functions.py, product_class.py, lists.py

import string
import datetime
import statistics
import sys

#xlsx
import os
import glob
import csv
import openpyxl
import pandas as pd

#import object class
from product_class import Product

#import helper function
from ws_functions import build_data_set

#checks for arg list
if(len(sys.argv) < 2):
	print("Usage: Specify which list you would like to use on the command line.")
	print("T = T-Series Printers")
	print("P = P-Series Printers")
	print("PI = Paper and Ink")
	sys.exit()

data_dump = False
pi = False
if "T" in sys.argv[1]:
	from lists_t_series import *
elif "PI" in sys.argv[1]:
	pi = True
	data_dump = True
	from lists_paper_ink import *
elif "P" in sys.argv[1]:
	from lists_p_series import *
else:
	print("Usage: Specify which list you would like to use on the command line.")
	print("T = T-Series Printers")
	print("P = P-Series Printers")
	print("PI = Paper and Ink")
	sys.exit()


def append_category(f, product_set, category, green_words, red_words):
	#picks correct category
	for product in product_set:
		write = True

		if product.added:
			write = False

		for word in red_words:
			if word in product.name:
				write = False

		if write:
			for substring in green_words:
				if substring in product.name and not product.added:
					try:
						f.write(product.channel+","+product.country+","+product.website+","+product.company+","+category+",")
						f.write(product.name+","+product.id+","+product.price+","+product.shipping+"\n")
						product.added = True
					except:
						pass


def build_condensed_table(f, product_set):
	header = ",Website,"
	for i in sku_targets:
		header = header + i + ","
	f.write(header+"\n")
	for website in website_targets:
		#write country
		line = ""
		if "(CA)" in website:
			line += "CA,"
			website = website.replace(' (CA)','')
		else:
			if "\n" in website:
				pass
			else:
				line += "US,"
		line += website+","
		for sku in sku_targets:
			product_found = False
			for product in product_set:
				if product.id == sku and product.website == website and not product_found:
					line += product.price

					#add product to avg_price_hash
					num = True
					try:
						price = product.price
						try:
							price = price.split('.', 1)[0]
						except:
							pass
						try:
							price = price.replace('*','')
						except:
							pass

						price = price.replace('$','')
						price = int(price)
					except:
						num = False

					try:
						if num:
							avg_price_hash[sku].append(price)
					except:
						pass

					#check if lowest sales price is correct
					try:
						lsp = price_target_hash[sku]
						up = price_up_hash[sku]
						if product.country == "US" and product.website != "Epson":
							if price < lsp:
								line = line+" <"
							if price > up:
								line = line+" >"
					except:
						pass

					if "Free" in product.shipping:
						line += "*,"
					else:
						line += ","

					product_found = True

			if not product_found:
				line += ","

		f.write(line.encode("utf-8")+"\n")


def find_averages(f, product_set):
	median_list = "\n,Median:,"
	mean_list = "\n,Mean:,"
	for sku in avg_price_hash:
		try:
			price_array = avg_price_hash[sku]
			median = statistics.median(price_array)
			mean = statistics.mean(price_array)
			median = round(median, 2)
			mean = round(mean, 2)
		except:
			mean = ""
			median = ""

		median_list += str(median)+","
		mean_list += str(mean)+","

	f.write(median_list)
	f.write(mean_list+"\n")


def highlight_prices():
	#convert .csv file to .xlsv
	wb = openpyxl.Workbook()
	ws = wb.active

	with open("/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv", 'rt') as f:
		reader = csv.reader(f)
		for r, row in enumerate(reader, start=1):
			for c, val in enumerate(row, start=1):
				try:
					val = val.replace('$','')
					ws.cell(row=r, column=c).value = float(val)
				except:
					ws.cell(row=r, column=c).value = val

	now = str(datetime.date.today())

	if "T" in sys.argv[1]:
		now = now + "_T-Series"
	elif "PI" in sys.argv[1]:
		now = now + "_Paper_and_Ink"
	elif "P" in sys.argv[1]:
		now = now + "_P-Series"

	#remove file if exists
	if os.path.isfile("/volume1/web/Epson_WS_Web/data_sheets/"+now+".xlsx"):
		os.remove("/volume1/web/Epson_WS_Web/data_sheets/"+now+".xlsx")

	wb.save("/volume1/web/Epson_WS_Web/data_sheets/"+now+".xlsx")

	#change letters to highlights
	df = pd.read_excel("/volume1/web/Epson_WS_Web/data_sheets/"+now+".xlsx")
	#chagne header location
	df.columns = df.iloc[2]
	df.reindex(df.index.drop(2))
	#writes over converted file
	fname = "/volume1/web/Epson_WS_Web/data_sheets/"+now+".xlsx"
	writer = pd.ExcelWriter(fname, engine='xlsxwriter')
	df.to_excel(writer, sheet_name=now[11:], index=False, header=False)

	# get xlsxwriter objects
	workbook  = writer.book
	worksheet = writer.sheets[now[11:]]

	#highlights incorrect prices
	formatH = workbook.add_format({'bg_color':'#FFC7CE', 'font_color':'#000000'})
	formatL = workbook.add_format({'bg_color':'#C6EFCE','font_color': '#000000'})
	worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
											'criteria': 'containing',
											'value':'<',
											'format': formatH})
	worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
											'criteria': 'containing',
											'value':'>',
											'format': formatL})

	writer.save()




# ****MAIN**** #
def main():
	#remove file if exists
	if os.path.isfile("/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv"):
		os.remove("/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv")
	#opens connection
	f = open(r"/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv", "w+")

	#hash set for all product objects
	product_set = set()
	build_data_set(product_set)
	print(str(len(product_set))+" potential products gathered.\n")
	#Writes name
	f.write("\nCreated by Sam Guyette\n")

	#Writes date and time of when ran
	now = datetime.datetime.now()
	f.write("Date and time of script execution: "+now.strftime("%Y-%m-%d %H:%M:%S")+"\n\n")

	if not data_dump:
		#builds condensed table
		print("Building condensed table...\n")
		title1 = "Condensed Table\n*Free Shipping\n                            =  Price bellow lowest sales price"
		title2 = " permitted (LSPP): <\n                            =  Price above unilateral price: >\n\n"
		up_prices = ",UP:,"
		lsp = ",LSPP:,"
		for key in price_up_hash:
			up_prices = up_prices + str(price_up_hash[key]) + ","
		for key in price_target_hash:
			lsp = lsp + str(price_target_hash[key]) + ","

		f.write(title1)
		f.write(title2)
		f.write(up_prices+"\n")
		f.write(lsp+"\n\n")
		build_condensed_table(f, product_set)
		find_averages(f, product_set)
		space = "\n\n\n\n\n"
		f.write(space)

	#write headers for super table
	title = "Super Table\n"
	f.write(title)
	headers = "Channel, Country, Website, Company, Category, Name, SKU, Price, Shipping\n"
	f.write(headers)

	#creates sub categories
	print("Writing all data to super table...\n")
	#builds super table
	try:
		append_category(f, product_set, "Printer", printer_include_list, printer_exclude_list)
		append_category(f, product_set, "Ink", ink_include_list, ink_exclude_list)
		append_category(f, product_set, "Accessory", accessories_include_list, accessories_exlude_list)
	except:
		print("Ink and paper only.")
		append_category(f, product_set, "Ink", ink_include_list, ink_exclude_list)
		append_category(f, product_set, "Paper", paper_include_list, paper_exlude_list)

	if not data_dump:
		#highlights prices that are over or under recommended selling point
		print("Highlighting prices...\n")
		highlight_prices()

		#remove csv file
		print("Deleting csv file...\n")
		os.remove("/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv")

	if pi:
		os.rename("/volume1/web/Epson_WS_Web/data_sheets/rough_output.csv", "/volume1/web/Epson_WS_Web/data_sheets/"+str(datetime.date.today())+"_Printer_and_Ink_Data.csv")

	print("Intern work is now complete.")
	f.close()


if __name__ == '__main__':
	main()













#
